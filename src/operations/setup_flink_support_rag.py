#!/usr/bin/env python3
"""Set up a Vector Search index over the Flink support knowledge base volume.

Creates (or reuses) a Vector Search endpoint and a direct-access index backed
by documents in /Volumes/quickstart_catalog/multi_agent_schema/support_kb.

Usage:
    uv run assistant-setup-flink [--profile PROFILE] [--endpoint-name NAME]
"""

import argparse
import hashlib
import io
import os
import sys
import time
from pathlib import Path

from databricks.sdk import WorkspaceClient
from databricks.sdk.service.vectorsearch import (
    DeltaSyncVectorIndexSpecRequest,
    EmbeddingSourceColumn,
    EndpointType,
    PipelineType,
    VectorIndexType,
)
from docx import Document as DocxDocument
from dotenv import load_dotenv
from openpyxl import load_workbook
from pypdf import PdfReader

load_dotenv(dotenv_path=Path(__file__).parent.parent.parent / ".env", override=True)

VOLUME_PATH = "/Volumes/quickstart_catalog/multi_agent_schema/support_kb"
CATALOG = "quickstart_catalog"
SCHEMA = "multi_agent_schema"
TABLE_NAME = "flink_support_kb"
FULL_TABLE_NAME = f"{CATALOG}.{SCHEMA}.{TABLE_NAME}"
INDEX_NAME = f"{CATALOG}.{SCHEMA}.flink_support_search_index"
DEFAULT_ENDPOINT_NAME = "flink_support_ep"
EMBEDDING_MODEL = "databricks-gte-large-en"


def _wait_for_endpoint(w: WorkspaceClient, name: str, timeout: int = 600) -> None:
    """Poll until the vector search endpoint is online."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        ep = w.vector_search_endpoints.get_endpoint(name)
        status = getattr(ep, "endpoint_status", None)
        state = getattr(status, "state", None) if status else None
        if state and state.value == "ONLINE":
            print(f"  Endpoint {name!r} is ONLINE.")
            return
        print(f"  Endpoint status: {state} — waiting …")
        time.sleep(15)
    print(f"WARNING: endpoint {name!r} did not reach ONLINE within {timeout}s", file=sys.stderr)


def _wait_for_index(w: WorkspaceClient, index_name: str, timeout: int = 900) -> None:
    """Poll until the vector search index is ready."""
    deadline = time.time() + timeout
    while time.time() < deadline:
        idx = w.vector_search_indexes.get_index(index_name)
        status = getattr(idx, "status", None)
        ready = getattr(status, "ready", None) if status else None
        if ready:
            print(f"  Index {index_name!r} is READY.")
            return
        msg = getattr(status, "message", "") if status else ""
        print(f"  Index status: not ready ({msg}) — waiting …")
        time.sleep(20)
    print(f"WARNING: index {index_name!r} did not become ready within {timeout}s", file=sys.stderr)


def _extract_docx_text(data: bytes) -> str:
    """Extract plain text from a .docx file's bytes."""
    doc = DocxDocument(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_pdf_text(data: bytes) -> str:
    """Extract plain text from a PDF file's bytes."""
    reader = PdfReader(io.BytesIO(data))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_xlsx_text(data: bytes) -> str:
    """Extract cell values from an Excel workbook as text."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"[Sheet: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                lines.append(row_text)
    wb.close()
    return "\n".join(lines)


def _list_volume_files(w: WorkspaceClient, path: str) -> list[str]:
    """Recursively list file paths in a volume."""
    paths: list[str] = []
    for entry in w.files.list_directory_contents(path):
        if entry.is_directory:
            paths.extend(_list_volume_files(w, entry.path))
        else:
            paths.append(entry.path)
    return paths


def _ensure_source_table(w: WorkspaceClient) -> None:
    """Create or refresh the source Delta table from volume documents."""
    print(f"\n→ Building source table {FULL_TABLE_NAME} from {VOLUME_PATH} …")

    wh_id = _get_warehouse_id(w)

    # Create table with CDF enabled
    for stmt in [
        f"""
        CREATE TABLE IF NOT EXISTS {FULL_TABLE_NAME} (
            id STRING,
            content STRING,
            source STRING
        )
        USING DELTA
        TBLPROPERTIES (delta.enableChangeDataFeed = true)
        """,
        f"ALTER TABLE {FULL_TABLE_NAME} SET TBLPROPERTIES (delta.enableChangeDataFeed = true)",
        f"TRUNCATE TABLE {FULL_TABLE_NAME}",
    ]:
        w.statement_execution.execute_statement(
            warehouse_id=wh_id,
            statement=stmt.strip(),
            wait_timeout="50s",
        )

    # Read and parse documents from the volume
    files = _list_volume_files(w, VOLUME_PATH)
    rows_inserted = 0
    for file_path in files:
        ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""
        resp = w.files.download(file_path)
        raw = resp.contents.read()

        if ext == "docx":
            content = _extract_docx_text(raw)
        elif ext == "pdf":
            content = _extract_pdf_text(raw)
        elif ext == "xlsx":
            content = _extract_xlsx_text(raw)
        elif ext in ("txt", "md", "csv", "json", "log"):
            content = raw.decode("utf-8", errors="replace")
        else:
            # Skip unsupported binary formats
            print(f"  Skipping unsupported file: {file_path}")
            continue

        if not content.strip():
            continue

        doc_id = hashlib.md5(file_path.encode()).hexdigest()
        escaped_content = content.replace("'", "''")
        escaped_source = file_path.replace("'", "''")

        w.statement_execution.execute_statement(
            warehouse_id=wh_id,
            statement=f"INSERT INTO {FULL_TABLE_NAME} VALUES ('{doc_id}', '{escaped_content}', '{escaped_source}')",
            wait_timeout="50s",
        )
        rows_inserted += 1

    print(f"  Source table {FULL_TABLE_NAME} populated with {rows_inserted} documents.")


def _get_warehouse_id(w: WorkspaceClient) -> str:
    """Return a SQL warehouse ID from env or first available warehouse."""
    wh_id = os.getenv("UC_AUDIT_WAREHOUSE_ID", "").strip()
    if wh_id and wh_id != "unused-for-structured-logging":
        return wh_id
    warehouses = list(w.warehouses.list())
    if not warehouses:
        print("ERROR: No SQL warehouse found. Set UC_AUDIT_WAREHOUSE_ID.", file=sys.stderr)
        sys.exit(1)
    return warehouses[0].id


def _ensure_endpoint(w: WorkspaceClient, endpoint_name: str) -> None:
    """Create or reuse a vector search endpoint."""
    print(f"\n→ Ensuring vector search endpoint {endpoint_name!r} …")
    try:
        ep = w.vector_search_endpoints.get_endpoint(endpoint_name)
        print(f"  Endpoint {endpoint_name!r} already exists.")
        status = getattr(ep, "endpoint_status", None)
        state = getattr(status, "state", None) if status else None
        if state and state.value != "ONLINE":
            _wait_for_endpoint(w, endpoint_name)
    except Exception:
        print(f"  Creating endpoint {endpoint_name!r} …")
        w.vector_search_endpoints.create_endpoint(
            name=endpoint_name,
            endpoint_type=EndpointType.STANDARD,
        )
        _wait_for_endpoint(w, endpoint_name)


def _ensure_index(w: WorkspaceClient, endpoint_name: str) -> None:
    """Create or reuse a Delta Sync vector search index."""
    print(f"\n→ Ensuring vector search index {INDEX_NAME!r} …")
    try:
        w.vector_search_indexes.get_index(INDEX_NAME)
        print(f"  Index {INDEX_NAME!r} already exists.")
        return
    except Exception:
        pass

    print(f"  Creating Delta Sync index {INDEX_NAME!r} on endpoint {endpoint_name!r} …")
    w.vector_search_indexes.create_index(
        name=INDEX_NAME,
        endpoint_name=endpoint_name,
        primary_key="id",
        index_type=VectorIndexType.DELTA_SYNC,
        delta_sync_index_spec=DeltaSyncVectorIndexSpecRequest(
            source_table=FULL_TABLE_NAME,
            embedding_source_columns=[
                EmbeddingSourceColumn(
                    name="content",
                    embedding_model_endpoint_name=EMBEDDING_MODEL,
                )
            ],
            pipeline_type=PipelineType.TRIGGERED,
        ),
    )
    _wait_for_index(w, INDEX_NAME)


def main() -> None:
    parser = argparse.ArgumentParser(description="Set up Flink support RAG vector search index")
    parser.add_argument("--profile", default=None, help="Databricks CLI profile")
    parser.add_argument("--endpoint-name", default=DEFAULT_ENDPOINT_NAME, help="VS endpoint name")
    args = parser.parse_args()

    kwargs = {}
    if args.profile:
        kwargs["profile"] = args.profile
    w = WorkspaceClient(**kwargs)

    print("Setting up Flink Support RAG …")
    _ensure_source_table(w)
    _ensure_endpoint(w, args.endpoint_name)
    _ensure_index(w, args.endpoint_name)

    mcp_url = f"/api/2.0/mcp/ai-search/{CATALOG}/{SCHEMA}/flink_support_search_index"
    print("\n✓ Setup complete.")
    print(f"  MCP URL: {mcp_url}")
    print("  Add this MCP URL to your subagents.<target>.json configuration.")


if __name__ == "__main__":
    main()
