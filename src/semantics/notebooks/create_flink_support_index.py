# Databricks notebook source
# MAGIC %md
# MAGIC # Build flink_support_index
# MAGIC
# MAGIC Extracts text from the Flink support knowledge base volume, builds the
# MAGIC `flink_support_kb` source table, and creates/refreshes the `flink_support_index`
# MAGIC Vector Search index used by the `flink_support_agent` MCP tool.
# MAGIC
# MAGIC Run as a Databricks Job (see `resources/semantics_jobs.yml`). Safe to re-run.

# COMMAND ----------

# MAGIC %pip install -q python-docx pypdf openpyxl
# MAGIC dbutils.library.restartPython()

# COMMAND ----------

dbutils.widgets.text("catalog", "quickstart_catalog", "Unity Catalog catalog")
dbutils.widgets.text("schema", "multi_agent_schema", "Unity Catalog schema")
dbutils.widgets.text(
    "volume_path", "/Volumes/quickstart_catalog/multi_agent_schema/support_kb", "Support KB volume path"
)
dbutils.widgets.text("source_table", "flink_support_kb", "Source table for the index")
dbutils.widgets.text("index_name", "flink_support_index", "Vector Search index name")
dbutils.widgets.text("endpoint_name", "flink-support-vs-endpoint", "Vector Search endpoint name")
dbutils.widgets.text("embedding_model", "databricks-gte-large-en", "Embedding model endpoint")

catalog = dbutils.widgets.get("catalog")
schema = dbutils.widgets.get("schema")
volume_path = dbutils.widgets.get("volume_path")
source_table = dbutils.widgets.get("source_table")
index_name = dbutils.widgets.get("index_name")
endpoint_name = dbutils.widgets.get("endpoint_name")
embedding_model = dbutils.widgets.get("embedding_model")

full_source_table = f"{catalog}.{schema}.{source_table}"
full_index_name = f"{catalog}.{schema}.{index_name}"

# COMMAND ----------

# MAGIC %md
# MAGIC ## 1. Extract documents from the support KB volume into a Delta table

# COMMAND ----------

import hashlib
import io

from databricks.sdk import WorkspaceClient
from docx import Document as DocxDocument
from openpyxl import load_workbook
from pypdf import PdfReader

w = WorkspaceClient()


def _extract_docx_text(data: bytes) -> str:
    doc = DocxDocument(io.BytesIO(data))
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _extract_pdf_text(data: bytes) -> str:
    reader = PdfReader(io.BytesIO(data))
    return "\n".join(page.extract_text() or "" for page in reader.pages)


def _extract_xlsx_text(data: bytes) -> str:
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    lines: list[str] = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"[Sheet: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            row_text = "\t".join(str(c) if c is not None else "" for c in row)
            if row_text.strip():
                lines.append(row_text)
    wb.close()
    return "\n".join(lines)


def _list_volume_files(path: str) -> list[str]:
    paths: list[str] = []
    for entry in w.files.list_directory_contents(path):
        if entry.is_directory:
            paths.extend(_list_volume_files(entry.path))
        else:
            paths.append(entry.path)
    return paths


rows = []
for file_path in _list_volume_files(volume_path):
    ext = file_path.rsplit(".", 1)[-1].lower() if "." in file_path else ""
    raw = w.files.download(file_path).contents.read()

    if ext == "docx":
        content = _extract_docx_text(raw)
    elif ext == "pdf":
        content = _extract_pdf_text(raw)
    elif ext == "xlsx":
        content = _extract_xlsx_text(raw)
    elif ext in ("txt", "md", "csv", "json", "log"):
        content = raw.decode("utf-8", errors="replace")
    else:
        print(f"Skipping unsupported file: {file_path}")
        continue

    if not content.strip():
        continue

    rows.append((hashlib.md5(file_path.encode()).hexdigest(), content, file_path))

print(f"Extracted {len(rows)} documents from {volume_path}.")

# COMMAND ----------

spark.sql(f"""
CREATE TABLE IF NOT EXISTS {full_source_table} (
    id STRING,
    content STRING,
    source STRING
)
USING DELTA
TBLPROPERTIES (delta.enableChangeDataFeed = true)
""")

if rows:
    df = spark.createDataFrame(rows, schema="id STRING, content STRING, source STRING")
    df.createOrReplaceTempView("_flink_support_kb_staged")
    spark.sql(f"""
    MERGE INTO {full_source_table} AS target
    USING _flink_support_kb_staged AS source
    ON target.id = source.id
    WHEN MATCHED THEN UPDATE SET *
    WHEN NOT MATCHED THEN INSERT *
    """)

print(f"{full_source_table} refreshed with {len(rows)} documents.")

# COMMAND ----------

# MAGIC %md
# MAGIC ## 2. Ensure the Vector Search endpoint and Delta Sync index

# COMMAND ----------

import time

from databricks.sdk.service.vectorsearch import (
    DeltaSyncVectorIndexSpecRequest,
    EmbeddingSourceColumn,
    EndpointType,
    PipelineType,
    VectorIndexType,
)


def _wait_for_endpoint(name: str, timeout: int = 600) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        ep = w.vector_search_endpoints.get_endpoint(name)
        state = getattr(getattr(ep, "endpoint_status", None), "state", None)
        if state and state.value == "ONLINE":
            print(f"Endpoint {name!r} is ONLINE.")
            return
        time.sleep(15)
    print(f"WARNING: endpoint {name!r} did not reach ONLINE within {timeout}s.")


try:
    w.vector_search_endpoints.get_endpoint(endpoint_name)
    print(f"Endpoint {endpoint_name!r} already exists.")
except Exception:
    print(f"Creating endpoint {endpoint_name!r} …")
    w.vector_search_endpoints.create_endpoint(name=endpoint_name, endpoint_type=EndpointType.STANDARD)
    _wait_for_endpoint(endpoint_name)

try:
    w.vector_search_indexes.get_index(full_index_name)
    print(f"Index {full_index_name!r} already exists; triggering sync.")
    w.vector_search_indexes.sync_index(full_index_name)
except Exception:
    print(f"Creating Delta Sync index {full_index_name!r} …")
    w.vector_search_indexes.create_index(
        name=full_index_name,
        endpoint_name=endpoint_name,
        primary_key="id",
        index_type=VectorIndexType.DELTA_SYNC,
        delta_sync_index_spec=DeltaSyncVectorIndexSpecRequest(
            source_table=full_source_table,
            embedding_source_columns=[
                EmbeddingSourceColumn(name="content", embedding_model_endpoint_name=embedding_model)
            ],
            pipeline_type=PipelineType.TRIGGERED,
        ),
    )

print(f"Done. Index target: {full_index_name}")
